"""
Orion Stats - Export Service
Generates Excel exports with multiple sheets.
"""
import io
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Optional, Any

from app.schemas.schemas import FilterCondition
from app.services.stats_service import calculate_descriptive_stats


# Styling
HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="A0D0FF", size=11)
CELL_FONT = Font(name="Calibri", color="E8F0F9", size=10)
TOTAL_FILL = PatternFill(start_color="0D1421", end_color="0D1421", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, color="4ADE80", size=10)
BORDER = Border(
    bottom=Side(style="thin", color="2A3F5F"),
)
BANNER_FILL = PatternFill(start_color="0D1421", end_color="0D1421", fill_type="solid")


def _style_header(ws, row, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def _resolve_logo_path() -> Optional[Path]:
    """Resolve ORION wordmark path for export branding."""
    here = Path(__file__).resolve()
    candidates = [
        here.parents[1] / "assets" / "orion-wordmark-only.png",  # backend/app/assets
        here.parents[3] / "frontend" / "public" / "orion-wordmark-only.png",
        here.parents[3] / "Logo Orion Insights sem fundo.png",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def _paint_banner(ws, start_row=1, end_row=6, end_col=16):
    """Paint a dark brand banner area in the worksheet."""
    for row in range(start_row, end_row + 1):
        ws.row_dimensions[row].height = 24
        for col in range(1, end_col + 1):
            ws.cell(row=row, column=col).fill = BANNER_FILL


def _insert_logo(ws, anchor_cell="A1", max_width=260):
    """Insert ORION wordmark if available."""
    logo_path = _resolve_logo_path()
    if not logo_path:
        return
    try:
        img = XLImage(str(logo_path))
        if img.width > max_width:
            ratio = max_width / float(img.width)
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
        ws.add_image(img, anchor_cell)
    except Exception:
        # Keep export resilient if image loading fails.
        return


def _auto_width(ws, max_width=28):
    """Auto-size worksheet columns with a max width cap."""
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, max_width)


def _write_simple_table(ws, headers: list[str], rows: list[list[Any]], start_row=1):
    """Write a generic table with default dark theme styling."""
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_idx, value=header)
    _style_header(ws, start_row, len(headers))

    for row_idx, row_values in enumerate(rows, start_row + 1):
        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.border = BORDER

    _auto_width(ws)
    return start_row + len(rows) + 1


def _write_stats_table(ws, stats_list, start_row=1):
    """Write statistics to a worksheet."""
    headers = [
        "Variavel", "N", "Ausentes", "% Ausentes", "Media", "Mediana", "Moda",
        "D. Padrao", "Variancia", "CV%", "SEM", "Min", "Max", "Amplitude",
        "Q1", "Q3", "IQR", "P5", "P10", "P90", "P95",
        "Assimetria", "Curtose", "IC Inf", "IC Sup", "Soma"
    ]

    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_idx, value=header)
    _style_header(ws, start_row, len(headers))

    for row_idx, stat in enumerate(stats_list, start_row + 1):
        values = [
            stat.name, stat.count, stat.missing_count, stat.missing_pct,
            stat.mean, stat.median, stat.mode, stat.std, stat.variance,
            stat.cv, stat.sem, stat.min, stat.max, stat.range,
            stat.q1, stat.q3, stat.iqr, stat.p5, stat.p10, stat.p90, stat.p95,
            stat.skewness, stat.kurtosis, stat.ci_lower, stat.ci_upper, stat.sum,
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.border = BORDER

    _auto_width(ws, max_width=20)

    return start_row + len(stats_list) + 1


def _group_summary_maps(group_summaries):
    sample_size_map = {}
    pct_total_map = {}
    if not group_summaries:
        return sample_size_map, pct_total_map
    for gs in group_summaries:
        sample_size_map[gs.group_key] = gs.sample_size
        pct_total_map[gs.group_key] = gs.pct_of_total
    return sample_size_map, pct_total_map


def _format_filter_summary(
    filters: Optional[list[FilterCondition]],
    columns_meta: dict[str, str],
    max_columns: int = 6,
) -> str:
    """Human-readable filter summary for executive metadata."""
    if not filters:
        return "Sem filtros ativos"

    chunks = []
    for cond in filters[:max_columns]:
        col_name = columns_meta.get(cond.col_key, cond.col_key)
        selected_count = len(cond.values) if cond.values is not None else 0
        chunks.append(f"{col_name} ({selected_count} valor(es))")

    if len(filters) > max_columns:
        chunks.append(f"+{len(filters) - max_columns} filtro(s)")

    return "; ".join(chunks)


def _build_group_report_rows(grouped, group_summaries):
    """One row per (group, variable) with comprehensive stats."""
    rows = []
    group_n_map, group_pct_map = _group_summary_maps(group_summaries)

    for group_key, group_stats in grouped.items():
        group_n = group_n_map.get(group_key)
        group_pct = group_pct_map.get(group_key)
        for stat in group_stats:
            rows.append([
                group_key,
                group_n,
                group_pct,
                stat.name,
                stat.count,
                stat.missing_count,
                stat.missing_pct,
                stat.mean,
                stat.median,
                stat.mode,
                stat.std,
                stat.variance,
                stat.cv,
                stat.sem,
                stat.min,
                stat.max,
                stat.range,
                stat.q1,
                stat.q3,
                stat.iqr,
                stat.p5,
                stat.p10,
                stat.p90,
                stat.p95,
                stat.skewness,
                stat.kurtosis,
                stat.ci_lower,
                stat.ci_upper,
                stat.sum,
            ])
    return rows


def _build_group_matrix_table(grouped, variables, columns_meta, group_summaries):
    """
    Matrix-style table:
    - rows: groups
    - columns: selected variables x key metrics
    """
    metric_defs = [
        ("mean", "Media"),
        ("std", "DP"),
        ("cv", "CV%"),
        ("median", "Mediana"),
        ("q1", "Q1"),
        ("q3", "Q3"),
        ("min", "Min"),
        ("max", "Max"),
        ("count", "N"),
    ]

    headers = ["Grupo", "N Grupo", "% Total"]
    for var in variables:
        var_name = columns_meta.get(var, var)
        for _, metric_label in metric_defs:
            headers.append(f"{var_name} | {metric_label}")

    group_n_map, group_pct_map = _group_summary_maps(group_summaries)
    rows = []
    for group_key, group_stats in grouped.items():
        stats_by_var = {s.col_key: s for s in group_stats}
        row = [group_key, group_n_map.get(group_key), group_pct_map.get(group_key)]
        for var in variables:
            stat = stats_by_var.get(var)
            for metric_key, _ in metric_defs:
                row.append(getattr(stat, metric_key, None) if stat else None)
        rows.append(row)

    return headers, rows


def _build_group_ranking_rows(grouped, variables, columns_meta):
    """
    Ranking table sorted by mean (desc) for each selected variable.
    """
    rows = []
    for var in variables:
        var_name = columns_meta.get(var, var)
        variable_entries = []
        for group_key, group_stats in grouped.items():
            stat = next((s for s in group_stats if s.col_key == var), None)
            if not stat or stat.mean is None:
                continue
            variable_entries.append((group_key, stat))

        variable_entries.sort(key=lambda x: x[1].mean if x[1].mean is not None else float("-inf"), reverse=True)

        for idx, (group_key, stat) in enumerate(variable_entries, 1):
            rows.append([
                var_name,
                idx,
                group_key,
                stat.mean,
                stat.std,
                stat.cv,
                stat.median,
                stat.min,
                stat.max,
                stat.range,
                stat.count,
                stat.ci_lower,
                stat.ci_upper,
            ])
    return rows


def _safe_val(value, default=-1e12):
    return value if value is not None else default


def _safe_float(value) -> Optional[float]:
    """Return finite float values only."""
    if value is None:
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    if pd.isna(numeric) or not pd.notna(numeric):
        return None
    if numeric == float("inf") or numeric == float("-inf"):
        return None
    return numeric


def _format_num(value: Optional[float], digits: int = 4) -> str:
    """Consistent number formatter for executive narrative text."""
    safe = _safe_float(value)
    if safe is None:
        return "-"
    return f"{safe:.{digits}f}"


def _get_primary_variable_rows(grouped, variable_key: str) -> list[dict[str, Any]]:
    """Collect grouped metrics for one variable."""
    rows = []
    if not grouped:
        return rows

    for group_key, group_stats in grouped.items():
        stat = next((s for s in group_stats if s.col_key == variable_key), None)
        if not stat:
            continue
        rows.append({
            "group": group_key,
            "mean": _safe_float(stat.mean),
            "std": _safe_float(stat.std),
            "cv": _safe_float(stat.cv),
            "count": int(stat.count) if stat.count is not None else 0,
            "min": _safe_float(stat.min),
            "max": _safe_float(stat.max),
            "range": _safe_float(stat.range),
            "median": _safe_float(stat.median),
        })

    return rows


def _write_table_block(ws, title: str, headers: list[str], rows: list[list[Any]], start_row: int) -> tuple[int, int]:
    """Write titled table block and return header/data end rows."""
    ws.cell(row=start_row, column=1, value=title).font = Font(name="Calibri", bold=True, size=12, color="A0D0FF")
    header_row = start_row + 1
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=header_row, column=col_idx, value=header)
    _style_header(ws, header_row, len(headers))

    for row_idx, row_values in enumerate(rows, header_row + 1):
        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.border = BORDER

    data_end_row = header_row + len(rows)
    _auto_width(ws, max_width=40)
    return header_row, data_end_row


def _add_bar_chart(
    ws,
    chart_title: str,
    y_title: str,
    header_row: int,
    data_start_row: int,
    data_end_row: int,
    category_col: int,
    value_col: int,
    anchor_cell: str,
):
    """Add a single-series bar chart from table range."""
    if data_end_row < data_start_row:
        return

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = chart_title
    chart.y_axis.title = y_title
    chart.x_axis.title = "Grupo"
    chart.height = 6.4
    chart.width = 10.8

    values = Reference(ws, min_col=value_col, min_row=header_row, max_row=data_end_row)
    categories = Reference(ws, min_col=category_col, min_row=data_start_row, max_row=data_end_row)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = "7CB7E8"
        chart.series[0].graphicalProperties.line.solidFill = "7CB7E8"

    ws.add_chart(chart, anchor_cell)


def _write_chart_notes(ws, start_row: int, title: str, lines: list[str]):
    """Write concise chart reading notes."""
    ws.cell(row=start_row, column=9, value=title).font = Font(name="Calibri", bold=True, size=11, color="A0D0FF")
    row = start_row + 1
    for line in lines:
        ws.cell(row=row, column=9, value=f"- {line}").font = Font(name="Calibri", size=10, color="C8D8EC")
        row += 1


def _build_executive_variable_rows(grouped, variables, columns_meta):
    """Executive summary rows with best/worst/variability leaders per variable."""
    rows = []
    if not grouped:
        return rows

    for var in variables:
        var_name = columns_meta.get(var, var)
        entries = []
        for group_key, group_stats in grouped.items():
            stat = next((s for s in group_stats if s.col_key == var), None)
            if not stat or stat.mean is None:
                continue
            entries.append((group_key, stat))

        if not entries:
            continue

        best = max(entries, key=lambda x: _safe_val(x[1].mean))
        worst = min(entries, key=lambda x: _safe_val(x[1].mean))
        highest_std = max(entries, key=lambda x: _safe_val(x[1].std))
        highest_cv = max(entries, key=lambda x: _safe_val(x[1].cv))
        largest_n = max(entries, key=lambda x: _safe_val(x[1].count, default=0))
        spread = (best[1].mean - worst[1].mean) if best[1].mean is not None and worst[1].mean is not None else None

        rows.append([
            var_name,
            best[0], best[1].mean,
            worst[0], worst[1].mean,
            spread,
            highest_std[0], highest_std[1].std,
            highest_cv[0], highest_cv[1].cv,
            largest_n[0], largest_n[1].count,
            len(entries),
        ])

    return rows


def _build_executive_narrative_rows(
    grouped,
    variables: list[str],
    columns_meta: dict[str, str],
    tests,
) -> list[list[Any]]:
    """Build executive narrative lines with evidence and action recommendation."""
    rows = []
    if not grouped:
        return rows

    test_by_variable = {}
    for test in tests or []:
        test_by_variable[test.variable] = test

    for variable in variables:
        variable_name = columns_meta.get(variable, variable)
        values = _get_primary_variable_rows(grouped, variable)
        values = [v for v in values if v["mean"] is not None]
        if not values:
            continue

        best = max(values, key=lambda x: x["mean"] if x["mean"] is not None else float("-inf"))
        worst = min(values, key=lambda x: x["mean"] if x["mean"] is not None else float("inf"))
        largest_n = max(values, key=lambda x: x["count"])
        cv_values = [v["cv"] for v in values if v["cv"] is not None]
        avg_cv = sum(cv_values) / len(cv_values) if cv_values else None
        volatile_group = max(values, key=lambda x: x["cv"] if x["cv"] is not None else float("-inf")) if cv_values else None

        spread = (best["mean"] - worst["mean"]) if best["mean"] is not None and worst["mean"] is not None else None
        spread_pct = None
        if spread is not None and worst["mean"] not in (None, 0):
            spread_pct = abs(spread / worst["mean"]) * 100

        if spread_pct is not None and spread_pct >= 60:
            priority = "Alta"
        elif spread_pct is not None and spread_pct >= 30:
            priority = "Media"
        elif avg_cv is not None and avg_cv >= 35:
            priority = "Media"
        else:
            priority = "Normal"

        if spread_pct is not None and spread_pct >= 60:
            implication = "Diferenca material entre grupos; revisar padrao operacional e mix."
        elif spread_pct is not None and spread_pct >= 30:
            implication = "Variacao relevante por grupo; manter monitoramento em ciclos curtos."
        else:
            implication = "Variacao controlada; foco em manutencao e ganho incremental."

        test = test_by_variable.get(variable)
        if test:
            test_message = (
                f"{test.test_name_display} com diferenca significativa (p={_format_num(test.p_value, 6)})."
                if test.significant
                else f"{test.test_name_display} sem diferenca estatistica (p={_format_num(test.p_value, 6)})."
            )
        else:
            test_message = "Teste comparativo nao disponivel para esta variavel."

        finding = (
            f"Maior media em {best['group']} ({_format_num(best['mean'], 4)}) "
            f"e menor em {worst['group']} ({_format_num(worst['mean'], 4)})."
        )
        evidence = (
            f"Amplitude={_format_num(spread, 4)}; "
            f"CV medio={_format_num(avg_cv, 2)}%; "
            f"Grupo com maior volume={largest_n['group']} (N={largest_n['count']})."
        )
        action = (
            f"Usar {best['group']} como benchmark; reduzir volatilidade em "
            f"{volatile_group['group'] if volatile_group else worst['group']}."
        )

        rows.append([
            variable_name,
            finding,
            evidence,
            test_message,
            implication,
            action,
            priority,
        ])

    return rows


def _write_narrative_sheet(ws, grouped, variables, columns_meta, tests):
    """Write executive narrative sheet with recommendations."""
    _paint_banner(ws, start_row=1, end_row=5, end_col=18)
    _insert_logo(ws, anchor_cell="A1", max_width=300)
    ws.cell(row=6, column=1, value="Narrativas Executivas Automatizadas").font = Font(
        name="Calibri", bold=True, size=14, color="A0D0FF"
    )
    ws.cell(row=7, column=1, value="Objetivo: transformar estatistica em recomendacao de decisao.").font = Font(
        name="Calibri", size=10, color="8BA3C0"
    )

    if not grouped:
        ws.cell(row=9, column=1, value="Nao ha agrupamento ativo. Ative 'Agrupar por' para gerar narrativas comparativas.").font = CELL_FONT
        _auto_width(ws, max_width=48)
        return

    rows = _build_executive_narrative_rows(grouped, variables, columns_meta, tests)
    headers = [
        "Variavel",
        "Achado Principal",
        "Evidencia Numerica",
        "Resultado Estatistico",
        "Leitura de Negocio",
        "Acao Recomendada",
        "Prioridade",
    ]
    header_row, data_end_row = _write_table_block(
        ws=ws,
        title="Resumo estrategico por variavel",
        headers=headers,
        rows=rows,
        start_row=9,
    )

    for row in range(header_row + 1, data_end_row + 1):
        for col in range(2, 7):
            ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)
        priority_cell = ws.cell(row=row, column=7)
        if priority_cell.value == "Alta":
            priority_cell.font = Font(name="Calibri", bold=True, color="F87171")
        elif priority_cell.value == "Media":
            priority_cell.font = Font(name="Calibri", bold=True, color="FBBF24")
        else:
            priority_cell.font = Font(name="Calibri", bold=True, color="4ADE80")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 36
    ws.column_dimensions["G"].width = 12


def _write_glossary_sheet(ws):
    """Write metric dictionary and interpretation guide."""
    _paint_banner(ws, start_row=1, end_row=4, end_col=14)
    _insert_logo(ws, anchor_cell="A1", max_width=260)
    ws.cell(row=5, column=1, value="Glossario Executivo de Metricas").font = Font(
        name="Calibri", bold=True, size=14, color="A0D0FF"
    )

    headers = ["Metrica", "Definicao", "Como interpretar", "Faixa orientativa"]
    rows = [
        ["Media", "Valor medio do grupo.", "Compara nivel central entre grupos.", "Maior valor nao implica melhor sem contexto de negocio."],
        ["Mediana", "Ponto central da distribuicao.", "Menos sensivel a extremos.", "Use junto com media para detectar assimetria."],
        ["DP", "Desvio padrao.", "Mede dispersao absoluta.", "DP alto indica heterogeneidade interna."],
        ["CV%", "DP / media * 100.", "Mede dispersao relativa.", "<15% baixo; 15-30% moderado; >30% alto."],
        ["Min/Max", "Extremos observados.", "Mostra limites praticos do grupo.", "Gap alto pode sugerir mix diferente."],
        ["Q1/Q3/IQR", "Quartis e intervalo interquartil.", "Faixa tipica sem extremos.", "IQR alto indica variacao estrutural."],
        ["IC 95%", "Intervalo de confianca da media.", "Faixa provavel da media real.", "Sobreposicao alta reduz evidencia de diferenca."],
        ["Assimetria", "Direcao da cauda.", "Positiva: cauda a direita; negativa: esquerda.", "Valores |assimetria| > 1 indicam forte assimetria."],
        ["Curtose", "Peso das caudas.", "Alta curtose indica mais extremos.", ">3 geralmente indica caudas pesadas."],
        ["p-valor", "Probabilidade sob H0.", "p <= 0.05 sugere diferenca estatistica.", "Sempre interpretar com tamanho de efeito."],
        ["Tamanho de efeito", "Magnitude pratica da diferenca.", "Complementa significancia estatistica.", "Priorize diferencas com efeito medio/grande."],
        ["N", "Tamanho amostral.", "Apoia confiabilidade das metricas.", "N muito baixo reduz estabilidade de inferencias."],
    ]
    _write_simple_table(ws, headers, rows, start_row=7)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 38
    for row in ws.iter_rows(min_row=8, max_row=8 + len(rows), min_col=2, max_col=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_dashboard_sheet(
    ws,
    grouped,
    variables: list[str],
    columns_meta: dict[str, str],
):
    """Write dashboard sheet with chart blocks and data interpretation."""
    _paint_banner(ws, start_row=1, end_row=5, end_col=18)
    _insert_logo(ws, anchor_cell="A1", max_width=300)
    ws.cell(row=6, column=1, value="Dashboard Executivo por Grupo").font = Font(
        name="Calibri", bold=True, size=14, color="A0D0FF"
    )

    if not grouped or not variables:
        ws.cell(row=8, column=1, value="Nao ha dados agrupados para montar o dashboard.").font = CELL_FONT
        _auto_width(ws, max_width=40)
        return

    primary_variable = variables[0]
    primary_name = columns_meta.get(primary_variable, primary_variable)
    base_rows = _get_primary_variable_rows(grouped, primary_variable)
    mean_rows = [r for r in base_rows if r["mean"] is not None]
    if not mean_rows:
        ws.cell(row=8, column=1, value="Nao ha medias validas para o agrupamento selecionado.").font = CELL_FONT
        _auto_width(ws, max_width=40)
        return

    ws.cell(row=7, column=1, value=f"Variavel de referencia dos graficos: {primary_name}").font = Font(
        name="Calibri", bold=True, size=10, color="8BA3C0"
    )
    ws.cell(row=8, column=1, value=f"Total de grupos com dado: {len(mean_rows)}").font = Font(
        name="Calibri", size=10, color="8BA3C0"
    )

    section_row = 10

    # Section 1: top means
    top_mean = sorted(mean_rows, key=lambda x: x["mean"], reverse=True)[:12]
    top_headers = ["Ranking", "Grupo", "Media", "DP", "CV%", "N"]
    top_rows = [
        [idx, row["group"], row["mean"], row["std"], row["cv"], row["count"]]
        for idx, row in enumerate(top_mean, 1)
    ]
    header_row, data_end_row = _write_table_block(
        ws,
        f"1) Top grupos por media - {primary_name}",
        top_headers,
        top_rows,
        start_row=section_row,
    )
    _add_bar_chart(
        ws=ws,
        chart_title=f"Top medias - {primary_name}",
        y_title="Media",
        header_row=header_row,
        data_start_row=header_row + 1,
        data_end_row=data_end_row,
        category_col=2,
        value_col=3,
        anchor_cell=f"I{section_row}",
    )
    top_best = top_mean[0]
    top_worst = top_mean[-1]
    _write_chart_notes(
        ws,
        start_row=section_row,
        title="Leitura do grafico",
        lines=[
            f"Maior media: {top_best['group']} = {_format_num(top_best['mean'], 4)}.",
            f"Menor media no top 12: {top_worst['group']} = {_format_num(top_worst['mean'], 4)}.",
            "Barras altas indicam maior nivel medio da variavel no grupo.",
        ],
    )
    section_row = max(section_row + 18, data_end_row + 4)

    # Section 2: highest variability
    cv_rows = [r for r in mean_rows if r["cv"] is not None and r["count"] >= 2]
    if cv_rows:
        top_cv = sorted(cv_rows, key=lambda x: x["cv"], reverse=True)[:12]
        cv_headers = ["Ranking", "Grupo", "CV%", "DP", "Media", "N"]
        cv_table_rows = [
            [idx, row["group"], row["cv"], row["std"], row["mean"], row["count"]]
            for idx, row in enumerate(top_cv, 1)
        ]
        cv_header_row, cv_data_end_row = _write_table_block(
            ws,
            f"2) Grupos com maior variabilidade relativa - {primary_name}",
            cv_headers,
            cv_table_rows,
            start_row=section_row,
        )
        _add_bar_chart(
            ws=ws,
            chart_title=f"Top CV% - {primary_name}",
            y_title="CV%",
            header_row=cv_header_row,
            data_start_row=cv_header_row + 1,
            data_end_row=cv_data_end_row,
            category_col=2,
            value_col=3,
            anchor_cell=f"I{section_row}",
        )
        _write_chart_notes(
            ws,
            start_row=section_row,
            title="Leitura do grafico",
            lines=[
                "CV% mede instabilidade relativa do grupo.",
                "Valores acima de 30% geralmente indicam alta variabilidade.",
                f"Maior CV identificado: {top_cv[0]['group']} = {_format_num(top_cv[0]['cv'], 2)}%.",
            ],
        )
        section_row = max(section_row + 18, cv_data_end_row + 4)

    # Section 3: largest sample
    top_n_rows = sorted(mean_rows, key=lambda x: x["count"], reverse=True)[:12]
    n_headers = ["Ranking", "Grupo", "N", "Media", "DP", "CV%"]
    n_table_rows = [
        [idx, row["group"], row["count"], row["mean"], row["std"], row["cv"]]
        for idx, row in enumerate(top_n_rows, 1)
    ]
    n_header_row, n_data_end_row = _write_table_block(
        ws,
        "3) Grupos com maior base amostral",
        n_headers,
        n_table_rows,
        start_row=section_row,
    )
    _add_bar_chart(
        ws=ws,
        chart_title="Top N por grupo",
        y_title="N",
        header_row=n_header_row,
        data_start_row=n_header_row + 1,
        data_end_row=n_data_end_row,
        category_col=2,
        value_col=3,
        anchor_cell=f"I{section_row}",
    )
    _write_chart_notes(
        ws,
        start_row=section_row,
        title="Leitura do grafico",
        lines=[
            "N maior aumenta robustez das comparacoes.",
            "Use grupos com N alto como referencia de estabilidade.",
            f"Grupo mais representativo: {top_n_rows[0]['group']} (N={top_n_rows[0]['count']}).",
        ],
    )
    section_row = max(section_row + 18, n_data_end_row + 4)

    # Section 4: spread by variable (when multiple variables selected)
    if len(variables) > 1:
        spread_rows = []
        for variable in variables:
            variable_name = columns_meta.get(variable, variable)
            var_rows = _get_primary_variable_rows(grouped, variable)
            var_rows = [r for r in var_rows if r["mean"] is not None]
            if not var_rows:
                continue
            best = max(var_rows, key=lambda x: x["mean"])
            worst = min(var_rows, key=lambda x: x["mean"])
            spread = best["mean"] - worst["mean"] if best["mean"] is not None and worst["mean"] is not None else None
            spread_rows.append([
                variable_name,
                best["group"],
                best["mean"],
                worst["group"],
                worst["mean"],
                spread,
                len(var_rows),
            ])

        if spread_rows:
            sp_headers = ["Variavel", "Melhor Grupo", "Media Melhor", "Pior Grupo", "Media Pior", "Amplitude", "Grupos c/ dado"]
            sp_header_row, sp_data_end_row = _write_table_block(
                ws,
                "4) Amplitude de medias entre grupos por variavel",
                sp_headers,
                spread_rows,
                start_row=section_row,
            )
            _add_bar_chart(
                ws=ws,
                chart_title="Amplitude de medias por variavel",
                y_title="Amplitude",
                header_row=sp_header_row,
                data_start_row=sp_header_row + 1,
                data_end_row=sp_data_end_row,
                category_col=1,
                value_col=6,
                anchor_cell=f"I{section_row}",
            )
            _write_chart_notes(
                ws,
                start_row=section_row,
                title="Leitura do grafico",
                lines=[
                    "Amplitude alta sinaliza heterogeneidade entre grupos.",
                    "Priorize variaveis com maior amplitude para padronizacao.",
                    f"Maior amplitude atual: {spread_rows[0][0]} = {_format_num(spread_rows[0][5], 4)}.",
                ],
            )

    _auto_width(ws, max_width=42)


def _write_executive_sheet(
    ws,
    sample_size: int,
    total_groups: Optional[int],
    variables: list[str],
    columns_meta: dict[str, str],
    grouped,
    group_by: Optional[list[str]],
    filters: Optional[list[FilterCondition]],
    tests,
):
    """Create executive cover sheet with strategic summary."""
    _paint_banner(ws, start_row=1, end_row=6, end_col=18)
    _insert_logo(ws, anchor_cell="A1", max_width=320)

    ws.cell(row=7, column=1, value="Relatorio Executivo - Estatisticas por Grupo").font = Font(
        name="Calibri", bold=True, size=16, color="A0D0FF"
    )
    ws.cell(row=8, column=1, value=f"Gerado em: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')}").font = Font(
        name="Calibri", size=10, color="8BA3C0"
    )

    summary_headers = ["Indicador", "Valor"]
    summary_rows = [
        ["Amostra total", sample_size],
        ["Total de grupos", total_groups if total_groups is not None else "-"],
        ["Variaveis analisadas", ", ".join(columns_meta.get(v, v) for v in variables) if variables else "-"],
        ["Agrupamento ativo", ", ".join(columns_meta.get(g, g) for g in group_by) if group_by else "Sem agrupamento"],
        ["Filtros ativos", _format_filter_summary(filters, columns_meta)],
        ["Testes comparativos", "Ativos" if tests else "Nao disponivel"],
        ["Tipo de relatorio", "Executivo automatizado com tabelas, rankings e dashboard grafico"],
    ]
    _write_simple_table(ws, summary_headers, summary_rows, start_row=10)

    if grouped:
        exec_headers = [
            "Variavel",
            "Melhor Grupo", "Media Melhor",
            "Pior Grupo", "Media Pior",
            "Amplitude Medias",
            "Maior DP (Grupo)", "DP",
            "Maior CV% (Grupo)", "CV%",
            "Maior N (Grupo)", "N",
            "Grupos com dado",
        ]
        exec_rows = _build_executive_variable_rows(grouped, variables, columns_meta)
        if exec_rows:
            next_row = _write_simple_table(ws, exec_headers, exec_rows, start_row=19)
        else:
            next_row = 20

        narrative_rows = _build_executive_narrative_rows(grouped, variables, columns_meta, tests)
        if narrative_rows:
            ws.cell(row=next_row + 1, column=1, value="Sintese executiva (acao orientada por dados)").font = Font(
                name="Calibri", bold=True, size=12, color="A0D0FF"
            )
            headers = ["Variavel", "Achado", "Acao recomendada", "Prioridade"]
            simple_rows = [[row[0], row[1], row[5], row[6]] for row in narrative_rows]
            _write_simple_table(ws, headers, simple_rows, start_row=next_row + 2)

    _auto_width(ws, max_width=42)


def create_excel_export(
    df: pd.DataFrame,
    variables: list[str],
    columns_meta: dict[str, str],
    filters: Optional[list[FilterCondition]] = None,
    group_by: Optional[list[str]] = None,
    treat_missing_as_zero: bool = True,
    include_sheets: list[str] = None,
) -> io.BytesIO:
    """Create Excel workbook with statistics."""
    if include_sheets is None:
        include_sheets = ["descriptive"]

    sample_size, stats, grouped, summaries, tests, total_groups = calculate_descriptive_stats(
        df=df, variables=variables, columns_meta=columns_meta,
        filters=filters, group_by=group_by,
        treat_missing_as_zero=treat_missing_as_zero,
        run_comparison_tests=True if group_by else False,
    )

    wb = Workbook()
    active_taken = False

    def _sheet(title: str):
        nonlocal active_taken
        if not active_taken:
            ws_local = wb.active
            ws_local.title = title
            active_taken = True
            return ws_local
        return wb.create_sheet(title)

    # Sheet 1: Executive report
    if "executive" in include_sheets:
        ws = _sheet("Relatorio Executivo")
        _write_executive_sheet(
            ws=ws,
            sample_size=sample_size,
            total_groups=total_groups,
            variables=variables,
            columns_meta=columns_meta,
            grouped=grouped,
            group_by=group_by,
            filters=filters,
            tests=tests,
        )
        if grouped:
            ws_dashboard = _sheet("Dashboard Executivo")
            _write_dashboard_sheet(
                ws=ws_dashboard,
                grouped=grouped,
                variables=variables,
                columns_meta=columns_meta,
            )
            ws_narrative = _sheet("Narrativas Exec")
            _write_narrative_sheet(
                ws=ws_narrative,
                grouped=grouped,
                variables=variables,
                columns_meta=columns_meta,
                tests=tests,
            )
        ws_glossary = _sheet("Glossario")
        _write_glossary_sheet(ws_glossary)

    # Sheet 2: Descriptive stats
    if "descriptive" in include_sheets:
        ws = _sheet("Estatisticas Gerais")
        _paint_banner(ws, start_row=1, end_row=5, end_col=16)
        _insert_logo(ws, anchor_cell="A1", max_width=260)
        ws.cell(row=6, column=1, value=f"Amostra: {sample_size} registros").font = Font(
            bold=True, size=12, color="A0D0FF"
        )
        _write_stats_table(ws, stats, start_row=8)

    # Sheet 3: Grouped stats
    if "grouped" in include_sheets and grouped:
        ws = _sheet("Por Grupo")
        current_row = 1
        for group_key, group_stats in grouped.items():
            ws.cell(row=current_row, column=1, value=f"Grupo: {group_key}").font = Font(bold=True, size=11, color="22D3EE")
            current_row = _write_stats_table(ws, group_stats, start_row=current_row + 1)
            current_row += 1

    # Sheet 4: Group report (one row per group x variable)
    if "group_report" in include_sheets and grouped:
        ws = _sheet("Resumo Grupo")
        headers = [
            "Grupo", "N Grupo", "% Total",
            "Variavel", "N", "Ausentes", "% Ausentes",
            "Media", "Mediana", "Moda", "DP", "Variancia", "CV%", "SEM",
            "Min", "Max", "Amplitude", "Q1", "Q3", "IQR",
            "P5", "P10", "P90", "P95", "Assimetria", "Curtose",
            "IC Inf", "IC Sup", "Soma"
        ]
        rows = _build_group_report_rows(grouped, summaries)
        _write_simple_table(ws, headers, rows, start_row=1)

    # Sheet 5: Group matrix (compact comparative view)
    if "group_matrix" in include_sheets and grouped:
        ws = _sheet("Matriz Grupo")
        headers, rows = _build_group_matrix_table(grouped, variables, columns_meta, summaries)
        _write_simple_table(ws, headers, rows, start_row=1)

    # Sheet 6: Mean rankings by variable
    if "group_ranking" in include_sheets and grouped:
        ws = _sheet("Ranking Medias")
        headers = [
            "Variavel", "Ranking", "Grupo",
            "Media", "DP", "CV%", "Mediana",
            "Min", "Max", "Amplitude", "N", "IC Inf", "IC Sup"
        ]
        rows = _build_group_ranking_rows(grouped, variables, columns_meta)
        _write_simple_table(ws, headers, rows, start_row=1)

    # Sheet 7: Comparison tests
    if "comparison" in include_sheets and tests:
        ws = _sheet("Testes Comparativos")
        headers = ["Variavel", "Teste", "Estatistica", "p-valor", "Significativo", "Tamanho Efeito", "Interpretacao"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        _style_header(ws, 1, len(headers))

        for row_idx, test in enumerate(tests, 2):
            values = [
                test.variable_name, test.test_name_display,
                test.statistic, test.p_value,
                "Sim" if test.significant else "Nao",
                f"{test.effect_size_name} = {test.effect_size}" if test.effect_size else "-",
                test.interpretation,
            ]
            for col_idx, v in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=v)
                cell.font = CELL_FONT

    # Remove default empty sheet if we created named ones
    if wb.sheetnames[0] == "Sheet" and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
